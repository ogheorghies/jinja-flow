"""Data source implementations: Excel, CSV, and file directories."""

from __future__ import annotations

import csv
import fnmatch
import json
from pathlib import Path
from typing import Any, Iterator

import yaml

from jflow.config import SourceDef


class FileResult:
    """Wraps a single file lookup result. Falsy when the file was not found."""

    __slots__ = ("_path", "_encoding", "_content", "_data_loaded", "_data")

    def __init__(self, path: Path | None, encoding: str = "utf-8") -> None:
        self._path = path
        self._encoding = encoding
        self._content: str | None = None
        self._data_loaded = False
        self._data: Any = None

    @property
    def exists(self) -> bool:
        return self._path is not None and self._path.is_file()

    @property
    def content(self) -> str:
        if self._content is None:
            if self.exists:
                self._content = self._path.read_text(encoding=self._encoding)
            else:
                self._content = ""
        return self._content

    @property
    def data(self) -> Any:
        if not self._data_loaded:
            self._data_loaded = True
            if self.exists:
                suffix = self._path.suffix.lower()
                if suffix == ".json":
                    self._data = json.loads(self.content)
                elif suffix in (".yaml", ".yml"):
                    self._data = yaml.safe_load(self.content)
                else:
                    self._data = None
            else:
                self._data = None
        return self._data

    @property
    def name(self) -> str:
        return self._path.name if self._path else ""

    @property
    def stem(self) -> str:
        return self._path.stem if self._path else ""

    @property
    def path(self) -> Path | None:
        return self._path if self.exists else None

    def __bool__(self) -> bool:
        return self.exists

    def __str__(self) -> str:
        return self.content


class FileSource:
    """Wraps a directory for file lookups via glob patterns."""

    def __init__(self, base: Path, recursive: bool = False, encoding: str = "utf-8"):
        self._base = base
        self._recursive = recursive
        self._encoding = encoding

    def _glob(self, pattern: str) -> list[Path]:
        if self._recursive:
            matches = list(self._base.rglob(pattern))
        else:
            matches = list(self._base.glob(pattern))
        return sorted(matches)

    def find(self, pattern: str) -> FileResult:
        matches = self._glob(pattern)
        if matches:
            return FileResult(matches[0], self._encoding)
        return FileResult(None, self._encoding)

    def list(self, pattern: str = "*") -> list[FileResult]:
        return [FileResult(p, self._encoding) for p in self._glob(pattern)]


class _RowDict(dict):
    """Dict subclass allowing attribute access for row fields."""

    def __getattr__(self, key: str) -> Any:
        try:
            return self[key]
        except KeyError:
            raise AttributeError(key)


class TabularSource:
    """Base for Excel and CSV sources — iterable of row dicts with .lookup()."""

    def __init__(self, rows: list[dict[str, Any]], headers: list[str]):
        self._rows = [_RowDict(r) for r in rows]
        self._headers = headers

    @property
    def rows(self) -> list[dict[str, Any]]:
        return self._rows

    @property
    def headers(self) -> list[str]:
        return self._headers

    def lookup(self, key: str, value: Any) -> dict[str, Any]:
        for row in self._rows:
            if row.get(key) == value:
                return row
        return _RowDict()

    def __iter__(self) -> Iterator[dict[str, Any]]:
        return iter(self._rows)

    def __len__(self) -> int:
        return len(self._rows)


class ExcelSource(TabularSource):
    """Loads an Excel worksheet as row dicts."""

    def __init__(self, file: Path, sheet: str | None = None):
        from openpyxl import load_workbook

        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        row_iter = ws.iter_rows(values_only=True)
        headers = [str(h) for h in next(row_iter)]
        rows = [dict(zip(headers, vals)) for vals in row_iter]
        wb.close()
        super().__init__(rows, headers)


class CsvSource(TabularSource):
    """Loads a CSV file as row dicts."""

    def __init__(
        self, file: Path, delimiter: str = ",", encoding: str = "utf-8"
    ):
        with open(file, newline="", encoding=encoding) as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            headers = reader.fieldnames or []
            rows = list(reader)
        super().__init__(rows, list(headers))


def build_source(sdef: SourceDef) -> FileSource | TabularSource:
    """Factory: instantiate the correct source from a SourceDef."""
    if sdef.type == "excel":
        return ExcelSource(sdef.file, sdef.sheet)
    elif sdef.type == "csv":
        return CsvSource(sdef.file, sdef.delimiter, sdef.encoding)
    elif sdef.type == "files":
        return FileSource(sdef.path, sdef.recursive, sdef.encoding)
    else:
        raise ValueError(f"Unknown source type: {sdef.type}")
